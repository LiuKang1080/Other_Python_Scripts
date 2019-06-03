import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('Book1.xlsx')

Sheet1 = wb.get_sheet_by_name('Sheet1')
#print(Sheet1.columns[1])

#ws = wb['Sheet1']

print(Sheet1.max_row)
print(get_column_letter(27))

#print(Sheet1.columns[1])
for x in Sheet1.columns:
    for y in x:
        print(y.value)